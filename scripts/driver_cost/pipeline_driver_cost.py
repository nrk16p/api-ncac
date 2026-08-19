"""
Monthly deliver-result pipeline — downloads ATMS batch-report Excel files
(type=monthly-driver-cost) and stores rows in MongoDB db `mena-bi`,
collection `driverCost`.

Memory-safe for small servers: Excel rows are streamed with openpyxl
read-only mode and inserted in batches — no pandas, no full-file DataFrame.

Usage:
    python3 pipeline_deliver_result.py           # update last 3 calendar months
    python3 pipeline_deliver_result.py --full    # all files of the current year
"""
import io
import logging
import os
import re
import sys
import warnings
from datetime import datetime, timezone
from pathlib import Path

import requests
import urllib3
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymongo import MongoClient

# Load .env from scripts/ directory (parent of deliver_result/)
load_dotenv(Path(__file__).parent.parent / ".env")

BASE_URL = "https://www.mena-atms.com"
USERNAME = os.getenv("ATMS_USERNAME")
PASSWORD = os.getenv("ATMS_PASSWORD")
MONGODB_URI = os.getenv("MONGODB_URI")

DB_NAME = "mena-bi"
COLLECTION = "driverCost"
RUNS_COLLECTION = "pipeline_runs"
SHEET_NAME = "driverCost"
BATCH_SIZE = 1000

warnings.simplefilter("ignore", urllib3.exceptions.InsecureRequestWarning)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")


def atms_login(session):
    session.post(
        f"{BASE_URL}/account/user/login",
        data={"username": USERNAME, "password": PASSWORD, "submit": "login", "next": ""},
        verify=False, timeout=30,
    ).raise_for_status()
    logging.info("ATMS login OK")


def list_files(session, year, month=None):
    """Return [{file_id, caption, branch, year, month}] from the CMS file index."""
    files, page = {}, 1
    while True:
        params = {
            "type": "monthly-driver-cost", "use_as": "batch-report",
            "ref_id": 1, "year": year, "month": month or "",
            "search_fields": "year,month", "order_by": "f.seq desc", "page": page,
        }
        r = session.get(f"{BASE_URL}/cms/file/index", params=params, verify=False, timeout=60)
        r.raise_for_status()
        table = re.search(r"<table.*?</table>", r.text, re.S)
        if not table:
            break
        found_new = False
        for row in re.findall(r"<tr.*?</tr>", table.group(0), re.S):
            dl = re.search(r'href="/cms/file/download/id/(\d+)"', row)
            if not dl:
                continue
            file_id = int(dl.group(1))
            if file_id in files:
                continue
            text = re.sub(r"<[^>]+>", "|", row)
            caption_m = re.search(r"ค่าเที่ยว พจส\.\s*:\s*(.+?)\s+on\s+(\d{4})-(\d{2})", text)
            if not caption_m:
                continue
            branch = caption_m.group(1).strip()
            f_year, f_month = int(caption_m.group(2)), int(caption_m.group(3))
            files[file_id] = {
                "file_id": file_id,
                "caption": f"ค่าเที่ยว พจส. : {branch} on {f_year}-{f_month:02d}",
                "branch": branch, "year": f_year, "month": f_month,
            }
            found_new = True
        if not found_new:
            break
        page += 1
    return sorted(files.values(), key=lambda f: (f["year"], f["month"], f["branch"]))


def sanitize_headers(header_row):
    """Strip/dedupe header names; unnamed columns become col_{i}."""
    cols, seen = [], {}
    for i, c in enumerate(header_row):
        name = str(c).strip() if c is not None else ""
        if not name or name == "nan":
            name = f"col_{i}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        cols.append(name)
    return cols


def stream_file_to_mongo(content, meta, col):
    """Stream one xlsx into Mongo in batches. Returns rows inserted."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    synced_at = datetime.now(timezone.utc)
    headers, batch, inserted = None, [], 0

    # Row 1 = title, row 2 = column headers, rows 3+ = data
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            continue
        if idx == 2:
            headers = sanitize_headers(row)
            continue
        if all(v is None for v in row):
            continue
        doc = dict(zip(headers, row))
        ldt = doc.get("LDT")
        doc["_ldt_base"] = str(ldt).split("#")[0] if ldt is not None else None
        doc["_file_id"] = meta["file_id"]
        doc["_branch"] = meta["branch"]
        doc["_year"] = meta["year"]
        doc["_month"] = meta["month"]
        doc["_synced_at"] = synced_at
        batch.append(doc)
        if len(batch) >= BATCH_SIZE:
            col.insert_many(batch)
            inserted += len(batch)
            batch = []
    if batch:
        col.insert_many(batch)
        inserted += len(batch)
    wb.close()
    return inserted


def target_periods(full):
    """(year, month) pairs to sync: current year for --full, else last 3 months."""
    now = datetime.now()
    if full:
        return [(now.year, None)]
    periods = []
    y, m = now.year, now.month
    for _ in range(3):
        periods.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return periods


def main():
    full = "--full" in sys.argv
    started_at = datetime.now(timezone.utc)
    session = requests.Session()
    atms_login(session)

    files, seen_ids = [], set()
    for year, month in target_periods(full):
        for f in list_files(session, year, month):
            if f["file_id"] not in seen_ids:
                seen_ids.add(f["file_id"])
                files.append(f)
    logging.info(f"{len(files)} file(s) to sync ({'full year' if full else 'last 3 months'})")

    client = MongoClient(MONGODB_URI)
    col = client[DB_NAME][COLLECTION]
    col.create_index([("_year", 1), ("_month", 1), ("_branch", 1)])
    col.create_index("LDT")

    total, file_results = 0, []
    for f in files:
        r = session.get(f"{BASE_URL}/cms/file/download/id/{f['file_id']}", verify=False, timeout=120)
        r.raise_for_status()
        col.delete_many({"_branch": f["branch"], "_year": f["year"], "_month": f["month"]})
        n = stream_file_to_mongo(r.content, f, col)
        total += n
        file_results.append({"caption": f["caption"], "rows": n})
        logging.info(f"  {f['caption']}: {n} rows")

    finished_at = datetime.now(timezone.utc)
    client[DB_NAME][RUNS_COLLECTION].insert_one({
        "pipeline": "driver_cost",
        "mode": "full" if full else "last_3_months",
        "files": file_results,
        "total_rows": total,
        "created_at": finished_at,
        "duration_s": round((finished_at - started_at).total_seconds(), 1),
        "status": "success",
    })
    logging.info(f"done — {total} rows across {len(files)} files → {DB_NAME}.{COLLECTION}")
    client.close()


if __name__ == "__main__":
    main()
